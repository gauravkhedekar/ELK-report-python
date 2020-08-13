import csv , subprocess
import pandas as pd
#import epoch_converter
import copy_to_main
import openpyxl
from openpyxl import load_workbook
import numpy as np
import gspread
from oauth2client.service_account import ServiceAccountCredentials
import google_api_python
import time

filename = "Logontype-10 with Adminstrator.csv"
filename1 = "second.csv"
fields = [] #includes rest of rows from original document
rows = []

with open(filename , "r") as csvfile:
    data = csv.reader(csvfile)
    fields = data.__next__()
    # print(fields)
    fields.insert(5,"User ID")
    # print(fields)
    for row in data:

        try:
            p = subprocess.Popen(["nslookup " + row[3]], stdout=subprocess.PIPE, shell=True)
            # print(p.communicate()[0].decode("utf-8").split("\n")[0].split("=")[1].split(".internal")[0])
            returned_data = p.communicate()[0].decode("utf-8").split("\n")[0].split("=")[1].split(".internal")[0]
            row.insert(5, returned_data)
            rows.append(row)
            # print(rows)
        except:
            pass
            row.insert(5, "not found")
            rows.append(row)



    print("total number of rows {}".format(data.line_num))

with open(filename1, "w" , newline='') as csvwrite:
     writer = csv.writer(csvwrite)
     writer.writerow(fields)
     writer.writerows(rows)

print(rows)

number=0
for row in rows:
    number=number+1
    print("api_number:{}".format(number))
    if (number%30)==0:
        time.sleep(100)
    print(str(row).strip("[]").split(","))
    google_api_python.copy_to_google_sheet(str(row).strip("[]").split(","), index=number+1)



#name = epoch_converter.start_date
name="dec"
df = pd.read_csv(r'second.csv')

df.to_excel(r'second_converted.xlsx',sheet_name=name , index=0)

###########################

with pd.ExcelWriter('second_converted'+name+'.xlsx') as writer1:
    df.to_excel(writer1 , sheet_name=name)

###########################

#copy_to_main.copy_to_main()

###########################
# book = load_workbook('second_converted.xlsx')
# book.create_sheet(name)

# writer.book = book
# book.get_sheet_by_name(name)
# read_file.to_excel(r'second_converted.xlsx',sheet_name=name)
# writer.save()
# writer.close()




